from __future__ import annotations

"""Excel 樣式與圖表模組 — 預設樣式定義與圖表生成工具。"""

from enum import Enum
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# 預設樣式
# ---------------------------------------------------------------------------

class StylePreset(str, Enum):
    """內建樣式預設名稱。"""

    HEADER = "header"
    DATA = "data"
    HIGHLIGHT = "highlight"
    ERROR = "error"
    SUCCESS = "success"
    TOTAL = "total"


_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

STYLE_MAP: Dict[str, Dict[str, Any]] = {
    StylePreset.HEADER: {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": _THIN_BORDER,
    },
    StylePreset.DATA: {
        "font": Font(size=10),
        "alignment": Alignment(vertical="center"),
        "border": _THIN_BORDER,
    },
    StylePreset.HIGHLIGHT: {
        "font": Font(bold=True, size=10),
        "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "border": _THIN_BORDER,
    },
    StylePreset.ERROR: {
        "font": Font(color="FF0000", size=10),
        "fill": PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
        "border": _THIN_BORDER,
    },
    StylePreset.SUCCESS: {
        "font": Font(color="008000", size=10),
        "fill": PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid"),
        "border": _THIN_BORDER,
    },
    StylePreset.TOTAL: {
        "font": Font(bold=True, size=11),
        "fill": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="double"),
            bottom=Side(style="double"),
        ),
    },
}


def apply_style(ws: Worksheet, row: int, col: int, preset: str) -> None:
    """
    將預設樣式套用到指定儲存格。

    Args:
        ws: 工作表物件。
        row: 行號（1-based）。
        col: 欄號（1-based）。
        preset: 樣式名稱，對應 STYLE_MAP 的 key。
    """
    style = STYLE_MAP.get(preset, STYLE_MAP.get(StylePreset.DATA, {}))
    cell = ws.cell(row=row, column=col)
    for attr, value in style.items():
        setattr(cell, attr, value)


def apply_style_to_range(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
    preset: str,
) -> None:
    """將預設樣式套用到一個矩形範圍。"""
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            apply_style(ws, r, c, preset)


# ---------------------------------------------------------------------------
# 圖表生成工具
# ---------------------------------------------------------------------------

class ChartType(str, Enum):
    """支援的圖表類型。"""

    BAR = "bar"
    LINE = "line"
    PIE = "pie"


def create_chart(
    ws: Worksheet,
    chart_type: ChartType,
    data_range: Tuple[int, int, int, int],
    categories_range: Optional[Tuple[int, int, int, int]] = None,
    title: str = "",
    x_axis_title: str = "",
    y_axis_title: str = "",
    position: str = "E2",
    width: float = 15.0,
    height: float = 10.0,
    series_titles: Optional[List[str]] = None,
) -> None:
    """
    在工作表上建立圖表。

    Args:
        ws: 目標工作表。
        chart_type: 圖表類型（bar / line / pie）。
        data_range: 資料範圍 (min_col, min_row, max_col, max_row)，1-based。
        categories_range: 分類軸範圍 (min_col, min_row, max_col, max_row)。
        title: 圖表標題。
        x_axis_title: X 軸標題。
        y_axis_title: Y 軸標題。
        position: 圖表放置位置（如 "E2"）。
        width: 圖表寬度（英吋）。
        height: 圖表高度（英吋）。
        series_titles: 各數據系列的標題清單。
    """
    chart_classes = {
        ChartType.BAR: BarChart,
        ChartType.LINE: LineChart,
        ChartType.PIE: PieChart,
    }
    chart_cls = chart_classes.get(chart_type, BarChart)
    chart = chart_cls()
    chart.title = title
    chart.width = width
    chart.height = height

    if chart_type != ChartType.PIE:
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title

    min_col, min_row, max_col, max_row = data_range
    data_ref = Reference(
        ws,
        min_col=min_col,
        min_row=min_row,
        max_col=max_col,
        max_row=max_row,
    )

    if categories_range:
        cat_min_col, cat_min_row, cat_max_col, cat_max_row = categories_range
        cats = Reference(
            ws,
            min_col=cat_min_col,
            min_row=cat_min_row,
            max_col=cat_max_col,
            max_row=cat_max_row,
        )
    else:
        cats = None

    if chart_type == ChartType.PIE:
        chart.add_data(data_ref, titles_from_data=True)
        if cats:
            chart.set_categories(cats)
    else:
        chart.add_data(data_ref, titles_from_data=True)
        if cats:
            chart.set_categories(cats)

    # 設定系列標題
    if series_titles:
        for i, s_title in enumerate(series_titles):
            if i < len(chart.series):
                chart.series[i].title = SeriesLabel(v=s_title)

    ws.add_chart(chart, position)
