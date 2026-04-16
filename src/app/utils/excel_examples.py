from __future__ import annotations

"""
ExcelHandler 使用範例 — 展示所有功能的完整程式碼。

直接執行：python -m app.utils.excel_examples
"""

import logging
from pathlib import Path

import pandas as pd

from app.utils.excel_handler import ExcelHandler
from app.utils.excel_styles import ChartType, StylePreset
from app.utils.excel_validation import ColumnRule, DataType

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

OUTPUT_DIR = Path("output_examples")
OUTPUT_DIR.mkdir(exist_ok=True)


def example_1_basic_write() -> None:
    """範例 1：基本寫入 — 將資料寫成美化 Excel。"""
    logger.info("=== 範例 1：基本寫入 ===")

    data = [
        {"姓名": "王小明", "部門": "製造", "分數": 92.5, "通過": True},
        {"姓名": "李大華", "部門": "品保", "分數": 88.0, "通過": True},
        {"姓名": "張美玲", "部門": "製造", "分數": 45.0, "通過": False},
        {"姓名": "陳志偉", "部門": "研發", "分數": 76.3, "通過": True},
    ]

    handler = ExcelHandler()
    path = handler.write(
        data=data,
        output_path=OUTPUT_DIR / "01_basic.xlsx",
        sheet_name="成績單",
        freeze_panes="A2",
        auto_filter=True,
    )
    logger.info("產出：%s", path)


def example_2_multiple_sheets() -> None:
    """範例 2：多 Sheet 寫入。"""
    logger.info("=== 範例 2：多 Sheet 寫入 ===")

    handler = ExcelHandler()
    path = handler.write_multiple_sheets(
        sheets={
            "Q1 業績": pd.DataFrame({
                "月份": ["一月", "二月", "三月"],
                "營收": [1200000, 1350000, 1100000],
            }),
            "Q2 業績": pd.DataFrame({
                "月份": ["四月", "五月", "六月"],
                "營收": [1400000, 1500000, 1250000],
            }),
        },
        output_path=OUTPUT_DIR / "02_multi_sheet.xlsx",
    )
    logger.info("產出：%s", path)


def example_3_validation() -> None:
    """範例 3：資料驗證。"""
    logger.info("=== 範例 3：資料驗證 ===")

    # 建立測試資料
    df = pd.DataFrame({
        "員工編號": ["E001", "E002", "", "E004", "E005"],
        "姓名": ["王小明", "李大華", "張美玲", None, "陳志偉"],
        "年齡": [28, 150, 35, 42, "abc"],
        "部門": ["製造", "品保", "銷售", "製造", "研發"],
        "email": [
            "wang@test.com",
            "not_an_email",
            "chang@test.com",
            "chen@test.com",
            "wei@test.com",
        ],
    })

    rules = [
        ColumnRule(
            name="員工編號",
            data_type=DataType.STRING,
            required=True,
            regex_pattern=r"^E\d{3}$",
        ),
        ColumnRule(
            name="姓名",
            data_type=DataType.STRING,
            required=True,
            min_length=2,
            max_length=20,
        ),
        ColumnRule(
            name="年齡",
            data_type=DataType.INTEGER,
            required=True,
            min_value=18,
            max_value=100,
        ),
        ColumnRule(
            name="部門",
            data_type=DataType.STRING,
            allowed_values=["製造", "品保", "研發", "管理"],
        ),
        ColumnRule(
            name="email",
            data_type=DataType.STRING,
            regex_pattern=r"^[\w\.\-]+@[\w\.\-]+\.\w+$",
        ),
    ]

    handler = ExcelHandler()
    result = handler.validate(df, rules)

    logger.info("驗證結果：%s", "通過" if result.is_valid else "失敗")
    logger.info("錯誤數：%d", result.error_count)
    for err in result.errors:
        logger.info(
            "  行 %d, 欄 '%s': [%s] %s",
            err.row,
            err.column,
            err.error_type,
            err.message,
        )
    logger.info("摘要：%s", result.summary())


def example_4_column_detection() -> None:
    """範例 4：自動偵測欄位資訊。"""
    logger.info("=== 範例 4：欄位偵測 ===")

    # 先寫一個測試檔
    df = pd.DataFrame({
        "產品代碼": ["A001", "A002", "A003"],
        "價格": [99.5, 150.0, 200.0],
        "數量": [10, 20, 30],
        "上架日期": pd.to_datetime(["2024-01-01", "2024-02-15", "2024-03-20"]),
        "備註": ["熱銷", None, "新品"],
    })
    test_file = OUTPUT_DIR / "04_detect_test.xlsx"
    ExcelHandler().write(df, output_path=test_file)

    # 偵測欄位
    handler = ExcelHandler(test_file)
    columns = handler.detect_columns()
    for col_info in columns:
        logger.info(
            "  欄位: %-10s | 型別: %-10s | 非空率: %5.1f%% | 範例: %s",
            col_info["name"],
            col_info["inferred_type"],
            col_info["non_null_rate"],
            col_info["sample_values"][:3],
        )


def example_5_auto_map_columns() -> None:
    """範例 5：自動欄位對應（用別名匹配）。"""
    logger.info("=== 範例 5：自動欄位對應 ===")

    df = pd.DataFrame({
        "Product Name": ["Widget A", "Widget B"],
        "Unit Price": [10.0, 20.0],
        "QTY": [100, 200],
    })

    rules = [
        ColumnRule(
            name="產品名稱",
            aliases=["Product Name", "product_name", "品名"],
        ),
        ColumnRule(
            name="單價",
            aliases=["Unit Price", "price", "Price"],
        ),
        ColumnRule(
            name="數量",
            aliases=["QTY", "Quantity", "qty"],
        ),
        ColumnRule(
            name="備註",
            aliases=["Remark", "Note"],
        ),
    ]

    handler = ExcelHandler()
    mapping = handler.auto_map_columns(rules, df=df)
    logger.info("對應結果：%s", mapping)
    # 預期：{'產品名稱': 'Product Name', '單價': 'Unit Price', '數量': 'QTY'}
    # '備註' 找不到，不會出現


def example_6_read_and_validate_pipeline() -> None:
    """範例 6：完整管線 — 讀取 → 偵測 → 對應 → 驗證 → 轉換。"""
    logger.info("=== 範例 6：完整管線 ===")

    # 1. 建立模擬檔案
    df_source = pd.DataFrame({
        "Board Name": ["PCB-001", "PCB-002", "PCB-003"],
        "Panel ID": ["P01", "P02", ""],
        "Test Value": [1.5, 2.0, 999.0],
    })
    test_file = OUTPUT_DIR / "06_pipeline_test.xlsx"
    ExcelHandler().write(df_source, output_path=test_file)

    # 2. 讀取
    handler = ExcelHandler(test_file)
    df = handler.read_sheet()
    logger.info("讀取 %d 行", len(df))

    # 3. 定義規則
    rules = [
        ColumnRule(
            name="board_name",
            aliases=["Board Name", "板子名稱"],
            data_type=DataType.STRING,
            required=True,
        ),
        ColumnRule(
            name="panel_id",
            aliases=["Panel ID", "面板編號"],
            data_type=DataType.STRING,
            required=True,
        ),
        ColumnRule(
            name="test_value",
            aliases=["Test Value", "測試值"],
            data_type=DataType.FLOAT,
            min_value=0,
            max_value=100,
        ),
    ]

    # 4. 自動對應
    mapping = handler.auto_map_columns(rules, df=df)
    logger.info("欄位對應：%s", mapping)

    # 5. 驗證
    result = handler.validate(df, rules, column_mapping=mapping)
    logger.info("驗證：%s（%d 錯誤）", result.is_valid, result.error_count)
    for err in result.errors:
        logger.info("  %s", err.message)

    # 6. 轉為 dict list（可直接入 DB）
    records = handler.to_dict_list(df)
    logger.info("轉換為 %d 筆記錄", len(records))


def example_7_chart() -> None:
    """範例 7：圖表生成。"""
    logger.info("=== 範例 7：圖表生成 ===")

    from openpyxl import Workbook
    from app.utils.excel_styles import apply_style_to_range

    wb = Workbook()
    ws = wb.active
    ws.title = "良率報表"

    # 寫入資料
    headers = ["批次", "良率 (%)", "不良數"]
    data_rows = [
        ["Batch-01", 98.5, 3],
        ["Batch-02", 95.2, 10],
        ["Batch-03", 99.1, 2],
        ["Batch-04", 92.0, 16],
        ["Batch-05", 97.8, 5],
    ]

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    apply_style_to_range(ws, 1, 1, 1, len(headers), StylePreset.HEADER)

    for r, row_data in enumerate(data_rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # 柱狀圖
    handler = ExcelHandler()
    handler.add_chart(
        ws,
        chart_type=ChartType.BAR,
        data_range=(2, 1, 2, 6),       # 良率欄 B1:B6（含標題）
        categories_range=(1, 2, 1, 6),  # 批次名 A2:A6
        title="各批次良率",
        position="E2",
        y_axis_title="良率 (%)",
    )

    # 折線圖
    handler.add_chart(
        ws,
        chart_type=ChartType.LINE,
        data_range=(3, 1, 3, 6),
        categories_range=(1, 2, 1, 6),
        title="不良數趨勢",
        position="E18",
        y_axis_title="不良數",
    )

    out = OUTPUT_DIR / "07_chart.xlsx"
    wb.save(str(out))
    wb.close()
    logger.info("產出：%s", out)


def example_8_merge_files() -> None:
    """範例 8：合併多個 Excel 檔案。"""
    logger.info("=== 範例 8：合併檔案 ===")

    # 建立模擬檔案
    for i in range(1, 4):
        ExcelHandler().write(
            data=pd.DataFrame({
                "項目": [f"Item-{i}-A", f"Item-{i}-B"],
                "數值": [i * 10, i * 20],
            }),
            output_path=OUTPUT_DIR / f"08_merge_source_{i}.xlsx",
        )

    # 合併
    handler = ExcelHandler()
    merged = handler.merge_files(
        file_paths=[
            OUTPUT_DIR / f"08_merge_source_{i}.xlsx" for i in range(1, 4)
        ],
        output_path=OUTPUT_DIR / "08_merged_result.xlsx",
    )
    logger.info("合併結果：\n%s", merged.to_string())


def example_9_bytesio_output() -> None:
    """範例 9：產出 BytesIO（適合 FastAPI StreamingResponse）。"""
    logger.info("=== 範例 9：BytesIO 輸出 ===")

    df = pd.DataFrame({"A": [1, 2, 3], "B": [4, 5, 6]})
    handler = ExcelHandler()
    buffer = handler.write(data=df, output_path=None)
    logger.info("BytesIO 大小：%d bytes", buffer.getbuffer().nbytes)

    # FastAPI 中可以這樣用：
    # from fastapi.responses import StreamingResponse
    # return StreamingResponse(
    #     buffer,
    #     media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    #     headers={"Content-Disposition": "attachment; filename=report.xlsx"},
    # )


if __name__ == "__main__":
    example_1_basic_write()
    example_2_multiple_sheets()
    example_3_validation()
    example_4_column_detection()
    example_5_auto_map_columns()
    example_6_read_and_validate_pipeline()
    example_7_chart()
    example_8_merge_files()
    example_9_bytesio_output()
    logger.info("所有範例執行完成！")
