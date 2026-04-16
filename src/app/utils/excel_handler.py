from __future__ import annotations

"""
ExcelHandler — 通用 Excel 資料處理類別。

功能涵蓋：
- 讀取：單/多 Sheet、自動偵測表頭、大檔案串流讀取
- 寫入：含格式樣式、自動欄寬、凍結窗格、圖表生成
- 驗證：型別檢查、必填欄位、數值範圍、正則、自訂規則
- 合併：多檔案 / 多 Sheet 合併
- 匯出：DataFrame → Excel（含美化）

用法範例請見 excel_examples.py。
"""

import logging
import os
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import (
    Any,
    BinaryIO,
    Callable,
    Dict,
    Iterator,
    List,
    Optional,
    Sequence,
    Set,
    Tuple,
    Union,
)

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.excel_styles import (
    ChartType,
    StylePreset,
    apply_style,
    apply_style_to_range,
    create_chart,
)
from app.utils.excel_validation import (
    ColumnRule,
    DataType,
    ValidationResult,
)

logger = logging.getLogger(__name__)


class ExcelHandler:
    """
    通用 Excel 資料處理器。

    設計原則：
    - 讀取操作回傳 pandas DataFrame，方便後續分析
    - 寫入操作接受 DataFrame 或 List[Dict]，輸出美化後的 .xlsx
    - 所有方法皆為同步（非 async），在 FastAPI 中可用 run_in_executor 呼叫

    Attributes:
        file_path: Excel 檔案路徑（讀取時使用）。
        workbook: openpyxl Workbook 物件（寫入時使用）。
    """

    def __init__(self, file_path: Optional[Union[str, Path]] = None) -> None:
        self.file_path: Optional[Path] = Path(file_path) if file_path else None
        self.workbook: Optional[Workbook] = None
        self._dataframes: Dict[str, pd.DataFrame] = {}

    # ======================================================================
    # 讀取
    # ======================================================================

    def read_sheet(
        self,
        sheet_name: Union[str, int] = 0,
        header_row: Optional[int] = None,
        skip_rows: Optional[List[int]] = None,
        use_cols: Optional[List[str]] = None,
        dtype: Optional[Dict[str, Any]] = None,
        na_values: Optional[List[str]] = None,
    ) -> pd.DataFrame:
        """
        讀取單一工作表。

        Args:
            sheet_name: 工作表名稱或索引（0-based）。
            header_row: 表頭所在行（0-based）。None 則自動偵測。
            skip_rows: 要跳過的行號清單。
            use_cols: 只讀取指定欄位名稱。
            dtype: 強制指定欄位型別。
            na_values: 額外視為 NaN 的值。

        Returns:
            DataFrame 包含該 Sheet 的資料。
        """
        self._ensure_file_path()
        if header_row is None:
            header_row = self._detect_header_row(sheet_name)
            logger.info("自動偵測表頭位於第 %d 行（0-based）", header_row)

        default_na = ["", "N/A", "n/a", "NA", "na", "-", "--"]
        all_na = list(set(default_na + (na_values or [])))

        df = pd.read_excel(
            self.file_path,
            sheet_name=sheet_name,
            header=header_row,
            skiprows=skip_rows,
            usecols=use_cols,
            dtype=dtype,
            na_values=all_na,
            engine="openpyxl",
        )
        # 清理欄位名稱：去除前後空白、換行
        df.columns = [
            str(c).strip().replace("\n", " ").replace("\r", "")
            for c in df.columns
        ]
        # 去除全空行
        df.dropna(how="all", inplace=True)
        df.reset_index(drop=True, inplace=True)

        name = sheet_name if isinstance(sheet_name, str) else str(sheet_name)
        self._dataframes[name] = df
        logger.info(
            "讀取 Sheet '%s'：%d 行 x %d 欄",
            sheet_name,
            len(df),
            len(df.columns),
        )
        return df

    def read_all_sheets(
        self,
        header_row: Optional[int] = None,
    ) -> Dict[str, pd.DataFrame]:
        """
        讀取所有工作表。

        Args:
            header_row: 表頭所在行（所有 Sheet 共用）。None 則逐一自動偵測。

        Returns:
            字典 {sheet_name: DataFrame}。
        """
        self._ensure_file_path()
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        result: Dict[str, pd.DataFrame] = {}
        for name in sheet_names:
            result[name] = self.read_sheet(
                sheet_name=name,
                header_row=header_row,
            )
        self._dataframes = result
        return result

    def read_large_file(
        self,
        sheet_name: Union[str, int] = 0,
        chunk_size: int = 10000,
        header_row: int = 0,
    ) -> Iterator[pd.DataFrame]:
        """
        串流讀取大型 Excel 檔案，逐批回傳 DataFrame。

        底層使用 openpyxl read_only 模式，記憶體佔用極低。

        Args:
            sheet_name: 工作表名稱或索引。
            chunk_size: 每批行數。
            header_row: 表頭所在行（0-based）。

        Yields:
            每批次的 DataFrame。
        """
        self._ensure_file_path()
        wb = load_workbook(
            self.file_path, read_only=True, data_only=True
        )
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        headers: Optional[List[str]] = None
        current_row = 0
        chunk: List[Dict[str, Any]] = []

        for row_values in rows_iter:
            if current_row < header_row:
                current_row += 1
                continue
            if current_row == header_row:
                headers = [
                    str(v).strip() if v is not None else f"Column_{i}"
                    for i, v in enumerate(row_values)
                ]
                current_row += 1
                continue

            if headers is None:
                break

            row_dict = dict(zip(headers, row_values))
            # 跳過全空行
            if all(v is None for v in row_values):
                current_row += 1
                continue
            chunk.append(row_dict)
            current_row += 1

            if len(chunk) >= chunk_size:
                yield pd.DataFrame(chunk)
                chunk = []

        if chunk:
            yield pd.DataFrame(chunk)

        wb.close()
        logger.info(
            "串流讀取完成：共 %d 行", current_row - header_row - 1
        )

    # ======================================================================
    # 自動偵測
    # ======================================================================

    def _detect_header_row(
        self,
        sheet_name: Union[str, int] = 0,
        max_scan_rows: int = 20,
    ) -> int:
        """
        自動偵測表頭所在行。

        策略：找第一行「非空字串欄位比例 >= 60%」且「無純數字行」的行。

        Args:
            sheet_name: 工作表名稱或索引。
            max_scan_rows: 最多掃描幾行。

        Returns:
            表頭行號（0-based）。
        """
        self._ensure_file_path()
        wb = load_workbook(
            self.file_path, read_only=True, data_only=True
        )
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]

        best_row = 0
        best_score = 0.0

        for row_idx, row in enumerate(ws.iter_rows(
            max_row=max_scan_rows, values_only=True
        )):
            if not row:
                continue
            total = len(row)
            non_empty = sum(1 for v in row if v is not None)
            if non_empty == 0:
                continue

            string_count = sum(
                1 for v in row
                if v is not None and isinstance(v, str) and not v.strip().replace(".", "").isdigit()
            )
            score = string_count / total

            if score > best_score:
                best_score = score
                best_row = row_idx

        wb.close()
        return best_row

    def detect_columns(
        self,
        sheet_name: Union[str, int] = 0,
        header_row: Optional[int] = None,
    ) -> List[Dict[str, Any]]:
        """
        偵測欄位資訊，回傳每個欄位的名稱、推斷型別、非空率。

        Args:
            sheet_name: 工作表名稱或索引。
            header_row: 表頭行號。None 則自動偵測。

        Returns:
            欄位資訊清單，每筆包含 name / inferred_type / non_null_rate / sample_values。
        """
        df = self.read_sheet(sheet_name=sheet_name, header_row=header_row)
        columns_info: List[Dict[str, Any]] = []

        for col in df.columns:
            series = df[col]
            non_null = series.notna().sum()
            total = len(series)
            non_null_rate = round(non_null / total * 100, 1) if total > 0 else 0.0

            inferred = self._infer_column_type(series)
            samples = (
                series.dropna().head(5).tolist() if non_null > 0 else []
            )

            columns_info.append({
                "name": col,
                "inferred_type": inferred,
                "non_null_rate": non_null_rate,
                "non_null_count": int(non_null),
                "total_count": total,
                "unique_count": int(series.nunique()),
                "sample_values": samples,
            })
        return columns_info

    def auto_map_columns(
        self,
        rules: List[ColumnRule],
        df: Optional[pd.DataFrame] = None,
        sheet_name: Union[str, int] = 0,
    ) -> Dict[str, str]:
        """
        根據 ColumnRule 的 name + aliases 自動比對 DataFrame 欄位。

        回傳 {rule.name: actual_column_name} 對應表。
        找不到的欄位不會出現在結果中。

        Args:
            rules: 欄位規則清單。
            df: 來源 DataFrame。若為 None 則從已讀取的 sheet 取。
            sheet_name: 若 df 為 None，使用哪個 Sheet。

        Returns:
            欄位對應字典。
        """
        if df is None:
            name_key = (
                sheet_name if isinstance(sheet_name, str) else str(sheet_name)
            )
            df = self._dataframes.get(name_key)
            if df is None:
                df = self.read_sheet(sheet_name=sheet_name)

        actual_columns = {c.strip().lower(): c for c in df.columns}
        mapping: Dict[str, str] = {}

        for rule in rules:
            candidates = [rule.name]
            if rule.aliases:
                candidates.extend(rule.aliases)

            for candidate in candidates:
                normalized = candidate.strip().lower()
                if normalized in actual_columns:
                    mapping[rule.name] = actual_columns[normalized]
                    break

        return mapping

    # ======================================================================
    # 驗證
    # ======================================================================

    def validate(
        self,
        df: pd.DataFrame,
        rules: List[ColumnRule],
        column_mapping: Optional[Dict[str, str]] = None,
        max_errors: int = 1000,
    ) -> ValidationResult:
        """
        根據 ColumnRule 清單驗證 DataFrame。

        Args:
            df: 要驗證的 DataFrame。
            rules: 欄位驗證規則清單。
            column_mapping: 欄位名稱對應表 {rule.name: actual_col}。
                            若為 None 則 rule.name 直接對應 df 欄位。
            max_errors: 最多收集幾筆錯誤（避免記憶體爆炸）。

        Returns:
            ValidationResult 物件。
        """
        result = ValidationResult(total_rows=len(df))
        mapping = column_mapping or {r.name: r.name for r in rules}

        for rule in rules:
            actual_col = mapping.get(rule.name)
            if actual_col is None or actual_col not in df.columns:
                if rule.required:
                    result.add_warning(
                        f"必填欄位 '{rule.name}' 找不到對應的 DataFrame 欄位"
                    )
                continue

            for idx, value in df[actual_col].items():
                if result.error_count >= max_errors:
                    result.add_warning(
                        f"錯誤數超過上限 {max_errors}，停止驗證"
                    )
                    return result

                row_num = int(idx) + 2  # +1 for 0-based, +1 for header
                self._validate_cell(result, rule, row_num, value)

        return result

    def _validate_cell(
        self,
        result: ValidationResult,
        rule: ColumnRule,
        row: int,
        value: Any,
    ) -> None:
        """驗證單一儲存格的值。"""
        is_empty = value is None or (isinstance(value, float) and pd.isna(value)) or str(value).strip() == ""

        # 必填檢查
        if rule.required and is_empty:
            result.add_error(
                row=row,
                column=rule.name,
                value=value,
                error_type="required",
                message=f"欄位 '{rule.name}' 為必填但值為空",
            )
            return

        if is_empty:
            return  # 非必填且為空，跳過後續檢查

        # 型別檢查
        converted = self._check_type(value, rule.data_type)
        if converted is None:
            result.add_error(
                row=row,
                column=rule.name,
                value=value,
                error_type="type",
                message=f"欄位 '{rule.name}' 預期 {rule.data_type.value}，但值為 '{value}'（{type(value).__name__}）",
            )
            return

        # 數值範圍檢查
        if rule.data_type in (DataType.INTEGER, DataType.FLOAT):
            num_val = float(converted)
            if rule.min_value is not None and num_val < rule.min_value:
                result.add_error(
                    row=row,
                    column=rule.name,
                    value=value,
                    error_type="min_value",
                    message=f"欄位 '{rule.name}' 值 {num_val} 小於最小值 {rule.min_value}",
                )
            if rule.max_value is not None and num_val > rule.max_value:
                result.add_error(
                    row=row,
                    column=rule.name,
                    value=value,
                    error_type="max_value",
                    message=f"欄位 '{rule.name}' 值 {num_val} 大於最大值 {rule.max_value}",
                )

        # 字串長度檢查
        if rule.data_type == DataType.STRING:
            str_val = str(converted)
            if rule.min_length is not None and len(str_val) < rule.min_length:
                result.add_error(
                    row=row,
                    column=rule.name,
                    value=value,
                    error_type="min_length",
                    message=f"欄位 '{rule.name}' 長度 {len(str_val)} 小於最短 {rule.min_length}",
                )
            if rule.max_length is not None and len(str_val) > rule.max_length:
                result.add_error(
                    row=row,
                    column=rule.name,
                    value=value,
                    error_type="max_length",
                    message=f"欄位 '{rule.name}' 長度 {len(str_val)} 超過最長 {rule.max_length}",
                )

        # 允許值清單檢查
        if rule.allowed_values is not None:
            if converted not in rule.allowed_values:
                result.add_error(
                    row=row,
                    column=rule.name,
                    value=value,
                    error_type="allowed_values",
                    message=f"欄位 '{rule.name}' 值 '{converted}' 不在允許清單 {rule.allowed_values} 中",
                )

        # 正則檢查
        if rule.regex_pattern is not None and rule.data_type == DataType.STRING:
            if not re.match(rule.regex_pattern, str(converted)):
                result.add_error(
                    row=row,
                    column=rule.name,
                    value=value,
                    error_type="regex",
                    message=f"欄位 '{rule.name}' 值 '{converted}' 不符合正則 '{rule.regex_pattern}'",
                )

        # 自訂驗證
        if rule.custom_validator is not None:
            is_valid, msg = rule.custom_validator(converted)
            if not is_valid:
                result.add_error(
                    row=row,
                    column=rule.name,
                    value=value,
                    error_type="custom",
                    message=f"欄位 '{rule.name}': {msg}",
                )

    # ======================================================================
    # 寫入
    # ======================================================================

    def write(
        self,
        data: Union[pd.DataFrame, List[Dict[str, Any]]],
        output_path: Optional[Union[str, Path]] = None,
        sheet_name: str = "Sheet1",
        style_header: bool = True,
        style_data: bool = True,
        auto_width: bool = True,
        freeze_panes: Optional[str] = "A2",
        auto_filter: bool = True,
    ) -> Union[Path, BytesIO]:
        """
        將資料寫入 Excel 檔案，附帶美化格式。

        Args:
            data: DataFrame 或 List[Dict] 格式的資料。
            output_path: 輸出路徑。None 則回傳 BytesIO（適合 HTTP response）。
            sheet_name: 工作表名稱。
            style_header: 是否套用表頭樣式。
            style_data: 是否套用資料欄樣式。
            auto_width: 是否自動調整欄寬。
            freeze_panes: 凍結窗格位置（如 "A2" 凍結第一行）。
            auto_filter: 是否啟用自動篩選。

        Returns:
            output_path 不為 None 時回傳 Path；否則回傳 BytesIO。
        """
        if isinstance(data, list):
            df = pd.DataFrame(data)
        else:
            df = data

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 寫入表頭 + 資料
        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=True), start=1
        ):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

        total_rows = len(df) + 1  # +1 for header
        total_cols = len(df.columns)

        # 套用樣式
        if style_header and total_cols > 0:
            apply_style_to_range(ws, 1, 1, 1, total_cols, StylePreset.HEADER)

        if style_data and total_rows > 1 and total_cols > 0:
            apply_style_to_range(
                ws, 2, 1, total_rows, total_cols, StylePreset.DATA
            )

        # 自動欄寬
        if auto_width:
            self._auto_adjust_width(ws)

        # 凍結窗格
        if freeze_panes:
            ws.freeze_panes = freeze_panes

        # 自動篩選
        if auto_filter and total_cols > 0:
            ws.auto_filter.ref = (
                f"A1:{get_column_letter(total_cols)}{total_rows}"
            )

        # 輸出
        if output_path is not None:
            out = Path(output_path)
            out.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(out))
            wb.close()
            logger.info("寫入完成：%s（%d 行）", out, len(df))
            return out
        else:
            buffer = BytesIO()
            wb.save(buffer)
            wb.close()
            buffer.seek(0)
            return buffer

    def write_multiple_sheets(
        self,
        sheets: Dict[str, Union[pd.DataFrame, List[Dict[str, Any]]]],
        output_path: Optional[Union[str, Path]] = None,
        style_header: bool = True,
        auto_width: bool = True,
    ) -> Union[Path, BytesIO]:
        """
        將多個 DataFrame 寫入同一個 Excel 的不同 Sheet。

        Args:
            sheets: {sheet_name: data} 字典。
            output_path: 輸出路徑。None 回傳 BytesIO。
            style_header: 是否套用表頭樣式。
            auto_width: 是否自動調整欄寬。

        Returns:
            Path 或 BytesIO。
        """
        wb = Workbook()
        # 移除預設空白 Sheet
        wb.remove(wb.active)

        for name, data in sheets.items():
            df = pd.DataFrame(data) if isinstance(data, list) else data
            ws = wb.create_sheet(title=name)

            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=True), start=1
            ):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            total_cols = len(df.columns)
            if style_header and total_cols > 0:
                apply_style_to_range(
                    ws, 1, 1, 1, total_cols, StylePreset.HEADER
                )

            if auto_width:
                self._auto_adjust_width(ws)

            ws.freeze_panes = "A2"

        if output_path is not None:
            out = Path(output_path)
            out.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(out))
            wb.close()
            return out
        else:
            buffer = BytesIO()
            wb.save(buffer)
            wb.close()
            buffer.seek(0)
            return buffer

    def add_chart(
        self,
        ws: Worksheet,
        chart_type: ChartType,
        data_range: Tuple[int, int, int, int],
        categories_range: Optional[Tuple[int, int, int, int]] = None,
        title: str = "",
        position: str = "E2",
        **kwargs: Any,
    ) -> None:
        """
        在指定工作表上新增圖表。

        Args:
            ws: 工作表物件。
            chart_type: 圖表類型。
            data_range: 資料範圍 (min_col, min_row, max_col, max_row)。
            categories_range: 分類軸範圍。
            title: 圖表標題。
            position: 圖表位置。
            **kwargs: 其它參數傳給 create_chart()。
        """
        create_chart(
            ws=ws,
            chart_type=chart_type,
            data_range=data_range,
            categories_range=categories_range,
            title=title,
            position=position,
            **kwargs,
        )

    # ======================================================================
    # 合併
    # ======================================================================

    def merge_files(
        self,
        file_paths: List[Union[str, Path]],
        output_path: Optional[Union[str, Path]] = None,
        sheet_name: Union[str, int] = 0,
        header_row: Optional[int] = None,
        add_source_column: bool = True,
    ) -> pd.DataFrame:
        """
        合併多個 Excel 檔案的同一 Sheet 為一個 DataFrame。

        Args:
            file_paths: Excel 檔案路徑清單。
            output_path: 若指定，合併結果同時寫入此路徑。
            sheet_name: 要合併的工作表名稱或索引。
            header_row: 表頭行號。None 則自動偵測。
            add_source_column: 是否新增 _source_file 欄位標示來源。

        Returns:
            合併後的 DataFrame。
        """
        frames: List[pd.DataFrame] = []
        for fp in file_paths:
            handler = ExcelHandler(fp)
            df = handler.read_sheet(
                sheet_name=sheet_name, header_row=header_row
            )
            if add_source_column:
                df["_source_file"] = Path(fp).name
            frames.append(df)

        merged = pd.concat(frames, ignore_index=True)
        logger.info(
            "合併 %d 個檔案：共 %d 行 x %d 欄",
            len(file_paths),
            len(merged),
            len(merged.columns),
        )

        if output_path:
            self.write(merged, output_path=output_path)

        return merged

    def merge_sheets(
        self,
        sheet_names: Optional[List[str]] = None,
        add_source_column: bool = True,
    ) -> pd.DataFrame:
        """
        合併同一檔案中多個 Sheet 為一個 DataFrame。

        Args:
            sheet_names: 要合併的 Sheet 名稱清單。None 則合併全部。
            add_source_column: 是否新增 _source_sheet 欄位。

        Returns:
            合併後的 DataFrame。
        """
        all_dfs = self.read_all_sheets()
        if sheet_names:
            all_dfs = {k: v for k, v in all_dfs.items() if k in sheet_names}

        frames: List[pd.DataFrame] = []
        for name, df in all_dfs.items():
            if add_source_column:
                df = df.copy()
                df["_source_sheet"] = name
            frames.append(df)

        merged = pd.concat(frames, ignore_index=True)
        logger.info(
            "合併 %d 個 Sheet：共 %d 行",
            len(frames),
            len(merged),
        )
        return merged

    # ======================================================================
    # 工具方法
    # ======================================================================

    def get_sheet_names(self) -> List[str]:
        """取得所有工作表名稱。"""
        self._ensure_file_path()
        wb = load_workbook(self.file_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names

    def get_sheet_info(self) -> List[Dict[str, Any]]:
        """
        取得每個 Sheet 的基本資訊（名稱、行數、欄數）。

        Returns:
            每個 Sheet 的資訊字典清單。
        """
        self._ensure_file_path()
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        info: List[Dict[str, Any]] = []
        for name in wb.sheetnames:
            ws = wb[name]
            info.append({
                "name": name,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
            })
        wb.close()
        return info

    def to_dict_list(
        self,
        df: Optional[pd.DataFrame] = None,
        sheet_name: Union[str, int] = 0,
    ) -> List[Dict[str, Any]]:
        """
        將 DataFrame 轉為 List[Dict] 格式。

        NaN 值會被轉為 None。
        """
        if df is None:
            name_key = (
                sheet_name if isinstance(sheet_name, str) else str(sheet_name)
            )
            df = self._dataframes.get(name_key)
            if df is None:
                raise ValueError(
                    f"找不到 Sheet '{sheet_name}' 的 DataFrame，請先呼叫 read_sheet()"
                )
        return df.where(df.notna(), None).to_dict(orient="records")

    # ======================================================================
    # 內部方法
    # ======================================================================

    def _ensure_file_path(self) -> None:
        """確認 file_path 已設定且檔案存在。"""
        if self.file_path is None:
            raise ValueError("file_path 未設定，請在建構時傳入或指定")
        if not self.file_path.exists():
            raise FileNotFoundError(f"找不到檔案：{self.file_path}")

    def _auto_adjust_width(
        self,
        ws: Worksheet,
        min_width: float = 8.0,
        max_width: float = 60.0,
        padding: float = 2.0,
    ) -> None:
        """自動調整欄寬，根據內容最大長度。"""
        for col_cells in ws.columns:
            max_len = 0.0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    # 中文字約佔 2 個字元寬
                    text = str(cell.value)
                    length = sum(
                        2 if ord(c) > 127 else 1 for c in text
                    )
                    if length > max_len:
                        max_len = length
            adjusted = min(max(max_len + padding, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted

    @staticmethod
    def _infer_column_type(series: pd.Series) -> str:
        """根據 pandas Series 推斷欄位型別。"""
        non_null = series.dropna()
        if len(non_null) == 0:
            return "unknown"

        dtype_str = str(series.dtype)
        if "int" in dtype_str:
            return DataType.INTEGER.value
        if "float" in dtype_str:
            # 檢查是否其實都是整數
            if all(v == int(v) for v in non_null if pd.notna(v)):
                return DataType.INTEGER.value
            return DataType.FLOAT.value
        if "bool" in dtype_str:
            return DataType.BOOLEAN.value
        if "datetime" in dtype_str:
            return DataType.DATETIME.value

        # object 型別嘗試進一步推斷
        sample = non_null.head(100)
        # 嘗試轉為數值
        numeric = pd.to_numeric(sample, errors="coerce")
        if numeric.notna().sum() / len(sample) > 0.8:
            if all(
                float(v) == int(float(v))
                for v in sample
                if pd.notna(pd.to_numeric(v, errors="coerce"))
            ):
                return DataType.INTEGER.value
            return DataType.FLOAT.value

        # 嘗試轉為日期
        try:
            dates = pd.to_datetime(sample, errors="coerce", infer_datetime_format=True)
            if dates.notna().sum() / len(sample) > 0.8:
                return DataType.DATETIME.value
        except Exception:
            pass

        return DataType.STRING.value

    @staticmethod
    def _check_type(value: Any, expected: DataType) -> Any:
        """
        嘗試將值轉型為預期型別。

        Returns:
            轉換後的值，或 None（轉換失敗）。
        """
        try:
            if expected == DataType.STRING:
                return str(value)
            elif expected == DataType.INTEGER:
                f = float(value)
                if f != int(f):
                    return None
                return int(f)
            elif expected == DataType.FLOAT:
                return float(value)
            elif expected == DataType.BOOLEAN:
                if isinstance(value, bool):
                    return value
                s = str(value).strip().lower()
                if s in ("true", "1", "yes", "y", "是"):
                    return True
                if s in ("false", "0", "no", "n", "否"):
                    return False
                return None
            elif expected in (DataType.DATE, DataType.DATETIME):
                if isinstance(value, datetime):
                    return value
                return pd.to_datetime(value)
            return value
        except (ValueError, TypeError, OverflowError):
            return None
